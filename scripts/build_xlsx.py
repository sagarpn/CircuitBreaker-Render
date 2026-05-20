#!/usr/bin/env python3
import json, sys, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

src_file, out_file = sys.argv[1], sys.argv[2]

with open(src_file) as f:
    rows = json.load(f)

wb = Workbook()

# ── Styles ────────────────────────────────────────────────
DARK   = "1A1A1A"
GOLD   = "B8860B"
WHITE  = "FFFFFF"
LGREY  = "F5F5F5"
CAT_COLORS = {
    'upper': 'C8D8F0', 'lower': 'C8F0D0',
    'hiit':  'F0D8C8', 'core':  'F0C8C8',
}
thin   = Side(style='thin', color='DDDDDD')
bord   = Border(left=thin, right=thin, top=thin, bottom=thin)

hdr_font  = Font(bold=True, color=WHITE, name='Arial', size=9)
hdr_fill  = PatternFill("solid", fgColor=DARK)
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def cell_style(ws, r, c, bold=False, fill=None, align='left', wrap=False):
    cell = ws.cell(r, c)
    cell.font      = Font(bold=bold, name='Arial', size=9)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border    = bord
    if fill: cell.fill = PatternFill("solid", fgColor=fill)
    return cell

# ── Helpers ───────────────────────────────────────────────
def is_burner(ex):
    t = ex.get('tags') or ''
    return 'burnout' in str(t)

def is_timed(ex):
    return ex.get('format') == 'timed' or is_burner(ex)

def get_eq(ex):
    eq = ex.get('equipment') or []
    if isinstance(eq, str):
        try: eq = json.loads(eq)
        except: eq = [e.strip() for e in eq.split(',') if e.strip()]
    return [str(e).lower() for e in eq if e and str(e).lower() != 'none']

def extract_reps_sets(reps_str):
    reps_str = (reps_str or '').strip()
    m = re.search(r'(\d+)\s+sets?\s*[x×]\s*(\d+)', reps_str, re.I)
    if m: return m.group(2), m.group(1)
    if re.search(r'to failure|max reps|max effort', reps_str, re.I): return 'to_failure', ''
    return reps_str.replace(re.sub(r'^\d+\s+sets?\s*[x×]\s*', '', reps_str), '').strip() or reps_str, ''

def yes_blank(val):
    if val in (True, 'yes', 'Yes', 'YES', 1, '1'): return 'yes'
    return ''

def get_reps_only(reps_str):
    reps_str = (reps_str or '').strip()
    if re.search(r'to failure|max reps|max effort', reps_str, re.I): return ''
    m = re.search(r'\d+\s+sets?\s*[x×]\s*(\d+)', reps_str, re.I)
    if m: return m.group(1)
    m2 = re.search(r'(\d+)', reps_str)
    return m2.group(1) if m2 else reps_str

def get_sets_only(reps_str):
    m = re.search(r'^(\d+)\s+sets?', (reps_str or '').strip(), re.I)
    return m.group(1) if m else ''

def get_max_timed(desc):
    m = re.search(r'(\d+)\s*(?:sec|second)', (desc or ''), re.I)
    return m.group(1) if m else ''

# ── SHEET 1: Exercises ────────────────────────────────────
ws = wb.active
ws.title = "Exercises"

COLS = [
    ('id',            6),
    ('name',          35),
    ('description',   50),
    ('category',      10),
    ('hiit',          6),
    ('strength',      9),
    ('core',          6),
    ('amrap',         7),
    ('lucky7',        8),
    ('compound',      10),
    ('burner',        8),
    ('core_burner',   12),
    ('hiit_burner',   12),
    ('unilateral',    11),
    ('plyometric',    11),
    ('bodyweight',    12),
    ('dumbbells',     11),
    ('bench',         7),
    ('reps',          8),
    ('sets',          6),
    ('to_failure',    10),
    ('timed',         7),
    ('max_reps_timed',14),
    ('muscle_group',  14),
    ('display_muscle',15),
    ('intensity',     10),
    ('slot_order',    11),
    ('flagged',       8),
]

headers = [c[0] for c in COLS]
widths  = [c[1] for c in COLS]

# Header row
for ci, h in enumerate(headers, 1):
    c = ws.cell(1, ci, h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = hdr_align; c.border = bord

ws.row_dimensions[1].height = 32
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = 'C2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'

# Data rows
for ri, ex in enumerate(rows, 2):
    cat  = (ex.get('category') or '').lower()
    eq   = get_eq(ex)
    burn = is_burner(ex)
    reps_str = ex.get('reps') or ''
    nm   = (ex.get('name') or '').lower()

    to_fail = 'yes' if re.search(r'to failure|max reps|max effort', reps_str, re.I) else ''
    timed   = 'yes' if is_timed(ex) else ''

    data = {
        'id':            str(ex.get('id') or ''),
        'name':          ex.get('name') or '',
        'description':   ex.get('description') or '',
        'category':      cat,
        'hiit':          'yes' if cat == 'hiit' else '',
        'strength':      'yes' if cat in ('upper','lower') else '',
        'core':          'yes' if cat == 'core' else '',
        'amrap':         yes_blank(ex.get('amrap')),
        'lucky7':        yes_blank(ex.get('lucky7')),
        'compound':      yes_blank(ex.get('is_compound')),
        'burner':        'yes' if burn else '',
        'core_burner':   'yes' if cat == 'core' and burn else '',
        'hiit_burner':   'yes' if cat == 'hiit' and burn else '',
        'unilateral':    'yes' if re.search(r'each side|each leg|each arm', reps_str, re.I) else '',
        'plyometric':    'yes' if any(k in nm for k in ['jump','bound','hop','tuck jump','star jump','broad jump']) else '',
        'bodyweight':    'yes' if (not eq or all(e == 'none' for e in eq)) else '',
        'dumbbells':     'yes' if 'dumbbells' in eq else '',
        'bench':         'yes' if 'bench' in eq else '',
        'reps':          get_reps_only(reps_str),
        'sets':          get_sets_only(reps_str),
        'to_failure':    to_fail,
        'timed':         timed,
        'max_reps_timed':get_max_timed(ex.get('description')),
        'muscle_group':  ex.get('muscle_group') or '',
        'display_muscle':ex.get('display_muscle') or '',
        'intensity':     str(ex.get('intensity') or ''),
        'slot_order':    str(ex.get('slot_order') or ex.get('ex_order') or ''),
        'flagged':       'yes' if ex.get('flagged') else '',
    }

    row_fill = CAT_COLORS.get(cat, 'FFFFFF') if ri % 2 == 0 else 'FFFFFF'
    for ci, h in enumerate(headers, 1):
        v = data.get(h, '')
        cell = ws.cell(ri, ci, v)
        cell.font      = Font(name='Arial', size=9)
        cell.border    = bord
        cell.alignment = Alignment(
            vertical='center',
            horizontal='center' if ci > 3 else 'left',
            wrap_text=(ci == 3)
        )
        if ci == 4:  # category — coloured
            cell.fill = PatternFill("solid", fgColor=CAT_COLORS.get(cat,'FFFFFF'))
            cell.font = Font(bold=True, name='Arial', size=9)
        elif h in ('hiit','strength','core','amrap','lucky7','compound','burner','core_burner',
                   'hiit_burner','unilateral','plyometric','bodyweight','dumbbells','bench',
                   'to_failure','timed','flagged') and v == 'yes':
            cell.font = Font(color='2E7D32', bold=True, name='Arial', size=9)

    ws.row_dimensions[ri].height = 18

# ── SHEET 2: Legend ───────────────────────────────────────
ls = wb.create_sheet("How to Fill")
ls.column_dimensions['A'].width = 18
ls.column_dimensions['B'].width = 55
ls.column_dimensions['C'].width = 40

legend_rows = [
    ("COLUMN", "ACCEPTED VALUES", "NOTES"),
    ("id", "Number 1-317", "Do not change — used for matching"),
    ("name", "Text", "Exercise name — used for matching on upload"),
    ("description", "Text", "How to perform the exercise"),
    ("category", "upper / lower / core / hiit", "Primary category"),
    ("hiit", "yes / leave blank", "Exercise is HIIT cardio"),
    ("strength", "yes / leave blank", "Exercise is upper or lower strength"),
    ("core", "yes / leave blank", "Exercise is a core exercise"),
    ("amrap", "yes / leave blank", "Eligible for AMRAP format workouts"),
    ("lucky7", "yes / leave blank", "Eligible for Lucky 7s format workouts"),
    ("compound", "yes / leave blank", "Multi-joint movement (press, squat, row)"),
    ("burner", "yes / leave blank", "Burnout finisher — always last in circuit"),
    ("core_burner", "yes / leave blank", "Core burnout finisher"),
    ("hiit_burner", "yes / leave blank", "HIIT burnout finisher"),
    ("unilateral", "yes / leave blank", "Single side — each arm or each leg"),
    ("plyometric", "yes / leave blank", "Jumping or explosive movement"),
    ("bodyweight", "yes / leave blank", "No equipment needed"),
    ("dumbbells", "yes / leave blank", "Requires dumbbells"),
    ("bench", "yes / leave blank", "Requires bench"),
    ("reps", "Number only — e.g. 12", "Rep count only, no sets prefix"),
    ("sets", "Number only — e.g. 3", "Number of sets"),
    ("to_failure", "yes / leave blank", "Do until failure — no fixed rep count"),
    ("timed", "yes / leave blank", "Time-based exercise (hold or seconds)"),
    ("max_reps_timed", "Seconds — e.g. 30", "Max reps in X seconds"),
    ("muscle_group", "chest / back / shoulders / biceps / triceps / quads / glutes / hamstrings / core", "Primary muscle group"),
    ("display_muscle", "Chest / Back / Shoulders / Biceps / Triceps / Quads / Glutes / Hamstrings / Stability / Abs / Obliques / Lower Abs / Full Body / Cardio / Agility / Power / Legs / Calves", "Label shown on exercise card"),
    ("intensity", "1 / 2 / 3 / 4 / 5", "1=very easy  2=easy  3=moderate  4=hard  5=max effort"),
    ("slot_order", "1 / 2 / 3", "Upper body only: 1=lead compound  2=secondary  3=isolation"),
    ("flagged", "yes / leave blank", "Admin flagged — hidden from all workouts"),
]

hf = Font(bold=True, color=WHITE, name='Arial', size=10)
hfill = PatternFill("solid", fgColor=DARK)

for ri2, (col, vals, notes) in enumerate(legend_rows, 1):
    a = ls.cell(ri2, 1, col)
    b = ls.cell(ri2, 2, vals)
    c = ls.cell(ri2, 3, notes)
    for cell in (a, b, c):
        cell.border    = bord
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.font      = Font(name='Arial', size=9)
    if ri2 == 1:
        for cell in (a, b, c):
            cell.font = hf; cell.fill = hfill
    ls.row_dimensions[ri2].height = 22

ls.freeze_panes = 'A2'

wb.save(out_file)
print(f"Saved {len(rows)} exercises to {out_file}")
